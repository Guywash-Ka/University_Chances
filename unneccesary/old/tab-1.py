import openpyxl

# Проверяет совместимость по пармаетру Pi
def pi_match(t):
    if ((t[0] == 1) and (t[1] == 0)) or ((t[0] == 0) and (t[1] == 1)):
        return False
    return True

# Проверяет параметрическую совместимость
def full_match(a, b):
    for i in range(33):
        if not pi_match((a[i], b[i])):
            return False
    return True

# Ищет совместимые слова к нужному слову
def find_alternatives(word):
    global params, words
    alternatives = []
    for j in params:
            if full_match(params[words.index(word)], j) and word != words[params.index(j)]:
                alternatives.append(words[params.index(j)])
    return alternatives


t = openpyxl.load_workbook('/home/vlad/Загрузки/Tab-1.xlsx')
sheet = t.get_sheet_by_name('Лист1')
words = []
params = []
for rows in range(5, 112):
    if (sheet.cell(row=rows, column=1).value != 'T1') and sheet.cell(row=rows, column=1).value:        
        words.append(sheet.cell(row=rows, column=1).value) 
        params.append([])
        for cell in range(6, 39):
            params[len(words)-1].append(sheet.cell(row=rows, column=cell).value)



chain1 = ['молодая', 'студент', 'читает', 'интересную', 'книга', 'на', 'диваном']
oldChain = chain1.copy()
for i in range(len(chain1)):
    chain1[i] = params[words.index(chain1[i])]
for i in range(6):
    if full_match(chain1[i], chain1[i+1]):
        print(f'{oldChain[i]} {oldChain[i+1]} | POSITIVE')
    else:
        #alternatives = []
        #for j in params:
           # if full_match(chain1[i], j) and oldChain[i] != words[params.index(j)]:
               # alternatives.append(words[params.index(j)])
        alternatives_for_first_word = find_alternatives(oldChain[i])
        alternatives_for_second_word = find_alternatives(oldChain[i])
        alternatives = alternatives_for_first_word + alternatives_for_second_word
        print(f'{oldChain[i]} {oldChain[i+1]} | Negative |  {alternatives}')


